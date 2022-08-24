import openpyxl
from parce_from_mygkh import parce



data_list_test = parce()
data_list = data_list_test[0]


def add_data_to_xlsx():

    # cоздаём xlsx фаил и вкладку
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('managing_companise')
    wb.save("managing_companise.xlsx")

    # открываем фаил и вкладку
    filename = 'managing_companise.xlsx'
    book = openpyxl.load_workbook(filename=filename)
    sheet = book['managing_companise']

    # создаём список из уникальных заголовков и создаём их в таблице
    headlines = []
    for mc in data_list:
        for keys in mc.keys():
            if keys in headlines:
                continue
            else:
                headlines.append(keys)
    print("headlines", headlines)
    sheet.append(headlines)

    # создаём список букв соответсвующих списку заголовков
    headlines_letter_list = []
    n1=0
    for row in sheet:
        for cell in row:
            headlines_letter_list.append(cell.column_letter)
    print('headlines_letter_list ', headlines_letter_list)

    # заполняем таблицу данными
    r = 2
    for mc in data_list:
        i1 = 0
        for i in range(len(headlines)):
            try:
                sheet[f'{headlines_letter_list[i]}{r}'] = mc[headlines[i]]

                i1=i+1
            except:
                continue
        r=r+1

    book.save(filename)
    print('add_data_to_xlsx done!')
